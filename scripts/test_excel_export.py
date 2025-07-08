#!/usr/bin/env python3
"""
测试任务结果Excel导出功能
"""

import requests
import json
import os

# 测试配置
BASE_URL = "http://localhost:5000/api"
TEST_USERNAME = "admin"
TEST_PASSWORD = "admin123"

def get_auth_token():
    """获取认证令牌"""
    login_url = f"{BASE_URL}/auth/login"
    data = {
        "username": TEST_USERNAME,
        "password": TEST_PASSWORD
    }
    
    try:
        response = requests.post(login_url, json=data)
        if response.status_code == 200:
            result = response.json()
            if result.get('success'):
                token = result.get('access_token')
                print(f"✅ 登录成功，获取token: {token[:20]}...")
                return token
            else:
                print(f"❌ 登录失败: {result.get('message')}")
                return None
        else:
            print(f"❌ 登录请求失败，状态码: {response.status_code}")
            return None
    except Exception as e:
        print(f"❌ 登录异常: {str(e)}")
        return None

def get_suitable_tasks(token):
    """获取适合Excel导出的任务列表"""
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }
    
    try:
        response = requests.get(f"{BASE_URL}/tasks", headers=headers)
        if response.status_code == 200:
            result = response.json()
            if result.get('success'):
                tasks = result.get('data', {}).get('tasks', [])
                
                # 筛选支持Excel导出的任务类型
                supported_types = ['standard_review', 'standard_recommendation', 'standard_compliance']
                suitable_tasks = [
                    task for task in tasks 
                    if task.get('task_type') in supported_types and task.get('status') == 'completed'
                ]
                
                print(f"✅ 找到 {len(suitable_tasks)} 个适合导出的任务")
                return suitable_tasks
            else:
                print(f"❌ 获取任务列表失败: {result.get('message')}")
                return []
        else:
            print(f"❌ 获取任务列表请求失败，状态码: {response.status_code}")
            return []
    except Exception as e:
        print(f"❌ 获取任务列表异常: {str(e)}")
        return []

def test_excel_export(token, task_id):
    """测试Excel导出功能"""
    headers = {
        'Authorization': f'Bearer {token}',
    }
    
    print(f"\n🔍 测试任务 {task_id} 的Excel导出")
    print("-" * 50)
    
    export_url = f"{BASE_URL}/tasks/{task_id}/results/export-excel"
    print(f"请求URL: {export_url}")
    
    try:
        response = requests.get(export_url, headers=headers)
        print(f"响应状态码: {response.status_code}")
        print(f"响应头Content-Type: {response.headers.get('content-type', '未设置')}")
        print(f"响应头Content-Disposition: {response.headers.get('content-disposition', '未设置')}")
        print(f"响应大小: {len(response.content)} 字节")
        
        if response.status_code == 200:
            print("✅ Excel导出请求成功")
            
            # 检查响应类型
            content_type = response.headers.get('content-type', '')
            if 'spreadsheetml' in content_type or 'excel' in content_type:
                print("✅ 响应类型为Excel文件")
                
                # 保存文件以验证
                filename = f"test_export_{task_id}.xlsx"
                with open(filename, 'wb') as f:
                    f.write(response.content)
                
                # 检查文件
                if os.path.exists(filename):
                    file_size = os.path.getsize(filename)
                    print(f"✅ 文件保存成功: {filename} ({file_size} 字节)")
                    
                    # 尝试验证Excel文件结构
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(filename)
                        ws = wb.active
                        print(f"✅ Excel文件结构验证成功")
                        print(f"   - 工作表名称: {ws.title}")
                        print(f"   - 数据行数: {ws.max_row}")
                        print(f"   - 数据列数: {ws.max_column}")
                        
                        # 显示前几行内容
                        print(f"   - 前几行内容:")
                        for row in range(1, min(8, ws.max_row + 1)):
                            row_data = []
                            for col in range(1, min(6, ws.max_column + 1)):
                                cell_value = ws.cell(row=row, column=col).value
                                row_data.append(str(cell_value)[:20] if cell_value else "")
                            print(f"     第{row}行: {' | '.join(row_data)}")
                            
                    except ImportError:
                        print("⚠️  无法验证Excel文件结构（缺少openpyxl库）")
                    except Exception as e:
                        print(f"⚠️  Excel文件结构验证失败: {str(e)}")
                    
                    # 清理测试文件
                    try:
                        os.remove(filename)
                        print(f"🧹 测试文件已清理: {filename}")
                    except:
                        print(f"⚠️  无法删除测试文件: {filename}")
                        
                else:
                    print("❌ 文件保存失败")
            else:
                print(f"⚠️  响应类型不是Excel文件: {content_type}")
                
        elif response.status_code == 400:
            try:
                error_data = response.json()
                print(f"⚠️  请求参数错误: {error_data.get('message', '未知错误')}")
            except:
                print(f"⚠️  请求参数错误: {response.text[:200]}")
                
        elif response.status_code == 404:
            print("⚠️  任务不存在")
            
        elif response.status_code == 403:
            print("⚠️  无权限访问")
            
        else:
            print(f"❌ Excel导出失败，状态码: {response.status_code}")
            try:
                error_data = response.json()
                print(f"   错误信息: {error_data.get('message', '未知错误')}")
            except:
                print(f"   响应内容: {response.text[:200]}")
                
    except Exception as e:
        print(f"❌ Excel导出请求异常: {str(e)}")

def main():
    """主函数"""
    print("="*60)
    print("任务结果Excel导出测试工具")
    print("="*60)
    
    # 获取认证令牌
    token = get_auth_token()
    if not token:
        print("❌ 无法获取认证令牌，测试终止")
        return
    
    # 获取适合的任务
    suitable_tasks = get_suitable_tasks(token)
    if not suitable_tasks:
        print("⚠️  没有找到适合Excel导出的任务")
        print("   支持的任务类型: standard_review, standard_recommendation, standard_compliance")
        print("   任务状态必须为: completed")
        return
    
    # 测试每个适合的任务
    print(f"\n📋 准备测试 {len(suitable_tasks)} 个任务的Excel导出功能")
    
    for i, task in enumerate(suitable_tasks[:3], 1):  # 限制测试前3个任务
        print(f"\n📄 任务 {i}/{min(len(suitable_tasks), 3)}")
        print(f"   ID: {task['id']}")
        print(f"   标题: {task['title']}")
        print(f"   类型: {task['task_type_display']} ({task['task_type']})")
        print(f"   状态: {task['status_display']} ({task['status']})")
        
        test_excel_export(token, task['id'])
    
    print(f"\n" + "="*60)
    print("测试完成")
    print("="*60)
    print("\n说明:")
    print("- 此测试验证Excel导出接口的功能和响应格式")
    print("- 仅测试已完成的支持分页查询的任务类型")
    print("- 会验证Excel文件的结构和内容格式")
    print("- 测试文件会自动清理")

if __name__ == "__main__":
    main() 